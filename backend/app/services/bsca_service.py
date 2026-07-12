"""
BSCEA Excel 填表服务（v2 - Python计算引擎版）
将功能点和项目参数填入 BSCEA 官方模板，所有计算在 Python 中完成
写入计算后的数值结果，不依赖 Excel 公式计算
"""
import os, json, re
from datetime import datetime
from copy import copy

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
TEMPLATE_PATH = os.path.join(BASE_DIR, "BSCEA官网下载小工具-软件开发费用评估模板（20251030发布版）.xlsx")
EXPORT_DIR = os.path.join(BASE_DIR, "backend", "exports")

# ============================================================
# 标准库：与 BSCEA 模板完全对齐
# ============================================================

# UFP 权重（预估功能点法 / 估算功能点法）
UFP_WEIGHTS_ESTIMATED = {"ILF": 35, "EIF": 15, "EI": None, "EO": None, "EQ": None}
UFP_WEIGHTS_DETAILED  = {"ILF": 10, "EIF": 7,  "EI": 4,  "EO": 5,  "EQ": 4}

# 复用程度调整系数
REUSE_FACTORS = {"high": 0.333, "medium": 0.667, "low": 1.0}
# 修改类型调整系数
MODIFY_FACTORS = {"new": 1.0, "modify": 0.8, "delete": 0.2}
# 中文映射
REUSE_LABELS = {"high": "高", "medium": "中", "low": "低"}
MODIFY_LABELS = {"new": "新增", "modify": "修改", "delete": "删除"}
COMPLEXITY_LABELS = {"low": "低", "medium": "中", "high": "高"}

# 规模变更调整因子
SCALE_TIMING_FACTORS = {
    "early_stage":  1.39,
    "mid_stage":    1.21,
    "late_stage":   1.10,
    "post_delivery": 1.00,
}

# 应用类型调整因子
APPLICATION_TYPE_FACTORS = {
    "业务处理": 1.0, "科技": 1.2, "多媒体": 1.3,
    "智能信息": 1.5, "基础软件/支撑软件": 1.7,
    "通信控制": 1.9, "流程控制": 2.0,
}

# 完整性级别因子
INTEGRITY_FACTORS = {
    "C/D级别或无明确完整性级别": 1.0,
    "没有明确的完整性级别或等级为C/D": 1.0,
    "B级别（基本完整）": 1.1,
    "完整性级别为A/B同时为达成完整性级别要求采取了特殊的设计及实现方式": 1.1,
    "A级别（完整）": 1.3,
    "完整性级别为A同时为达成完整性级别要求在软件开发全生命周期均采取了特定、明确的措施": 1.3,
}

# 开发语言因子
LANGUAGE_FACTORS = {
    "JAVA、C++、C#及其他同级别语言/平台": 1.0,
    "Visual Basic、Dephi、PowerBuilder及其他同级别语言/平台": 0.8,
    "Access、FoxPro及其他同级别语言/平台": 0.8,
    "C及其他同级别语言/平台": 1.2,
    "PowerBuilder、ASP及其他同级别语言/平台": 0.8,
}

# 团队背景因子
TEAM_FACTORS = {
    "为本行业开发过类似的项目": 0.8,
    "为其他行业开发过类似的项目，或为本行业开发过不同但相关的项目": 1.0,
    "没有为其他行业或其他项目开发过类似的项目": 1.2,
    "没有同类项目的背景": 1.2,
    "没有同类项目的背景": 1.2,
}

# 分布式/性能/可靠性/多重站点 评分→因子（用于非功能特征）
NF_SCORE_MAP = {0: "没有识别到任何需要分布式处理的情况", 1: "通过网络进行客户端/服务器及网络基础计算机系统分布处理和传输", 2: "通过网络和多个计算机在多个处理器上同时处理数据和系统"}
PERF_SCORE_MAP = {0: "没有提到或没有性能需求", 1: "应答时间或处理率对高峰时间或所有业务时间都很重要，对连动系统结束处理时间有限制", 2: "为满足性能要求，分析阶段即需设计性能分析任务或在设计阶段之后需使用设计工具"}
RELI_SCORE_MAP = {0: "发生故障时带来的不便可以忽略不计", 1: "发生故障时可轻易修复，带来一定不便或经济损失", 2: "发生故障时很难修复，带来很大不便或经济损失"}
MULTI_SCORE_MAP = {0: "没有多重站点的需求", 1: "在用途类似的硬件或软件环境下运行", 2: "在用途不同的硬件或软件环境下运行"}

# 人月基准单价（按城市）
CITY_PRICES = {
    "北京": {"dev": 3.2198, "ops": 2.6335},
    "上海": {"dev": 3.0, "ops": 2.5},
    "广州": {"dev": 2.8, "ops": 2.3},
    "深圳": {"dev": 3.1, "ops": 2.6},
}
DEFAULT_DEV_PRICE = 3.2198
DEFAULT_OPS_PRICE = 2.6335

# PDR 基准（2025年中国软件行业基准数据）
PDR_MEDIAN = 6.72  # 全行业中位数，人时/功能点


def _get_ufp(category: str, fp_method: str) -> int:
    """获取类别对应的 UFP 权重"""
    if fp_method == "estimated":
        return UFP_WEIGHTS_ESTIMATED.get(category, 0) or 0
    return UFP_WEIGHTS_DETAILED.get(category, 0) or 0


def _calc_us(ufp: int, complexity: str, reuse_level: str, modify_type: str) -> float:
    """计算 US（调整后的功能点规模）"""
    reuse_f = REUSE_FACTORS.get(reuse_level, 1.0)
    modify_f = MODIFY_FACTORS.get(modify_type, 1.0)
    return round(ufp * reuse_f * modify_f, 2)


def fill_excel(fp_items: list, params: dict, output_path: str = None) -> dict:
    """
    将功能点和参数填入BSCEA模板，Python完成所有计算，写入数值
    fp_items: [{category, description, fp_name, ufp, complexity, reuse_level, modify_type, ...}]
    params: {fp_method, scale_timing, application_type, city, ...}
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"BSCEA模板不存在: {TEMPLATE_PATH}")

    import openpyxl

    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # ── 解析参数 ─────────────────────────────────
    fp_method = params.get("fp_method", "estimated")  # estimated|detailed
    is_estimated = (fp_method == "estimated")
    scale_timing = params.get("scale_timing", "post_delivery")
    app_type = params.get("application_type", "业务处理")
    city = params.get("city", "北京")
    pdr_percentile = params.get("pdr_percentile", "median")  # lower|median|upper
    team = params.get("team_background", "为其他行业开发过类似的项目，或为本行业开发过不同但相关的项目")

    # 非功能特征评分
    dist_score = int(params.get("distributed_score", 0))
    perf_score = int(params.get("performance_score", 0))
    reli_score = int(params.get("reliability_score", 0))
    multi_score = int(params.get("multi_site_score", 0))

    # 人月单价
    city_price_map = CITY_PRICES.get(city, {"dev": DEFAULT_DEV_PRICE})
    unit_price = city_price_map.get("dev", DEFAULT_DEV_PRICE)
    if params.get("unit_price"):
        unit_price = float(params.get("unit_price"))

    # PDR
    if pdr_percentile == "lower":
        pdr = PDR_MEDIAN * 0.8
    elif pdr_percentile == "upper":
        pdr = PDR_MEDIAN * 1.2
    else:
        pdr = PDR_MEDIAN

    # 规模变更调整因子
    scale_factor = SCALE_TIMING_FACTORS.get(scale_timing, 1.0)

    # ── 1. 填写项目特征 ──────────────────────────
    ws_feature = wb["项目特征"]

    # 规模计数时机
    timing_labels = {
        "early_stage": "估算早期（概算）", "mid_stage": "估算中期（预算）",
        "late_stage": "估算晚期（招标）", "post_delivery": "项目交付后及运维阶段",
    }
    ws_feature["C1"] = timing_labels.get(scale_timing, "项目交付后及运维阶段")
    ws_feature["D1"] = scale_factor

    # 应用类型
    ws_feature["C2"] = app_type
    app_factor = APPLICATION_TYPE_FACTORS.get(app_type, 1.0)
    ws_feature["D2"] = app_factor

    # 非功能性特征 - 分布式处理
    ws_feature["C3"] = NF_SCORE_MAP.get(dist_score, NF_SCORE_MAP[0])
    ws_feature["D3"] = dist_score
    # 性能
    ws_feature["C4"] = PERF_SCORE_MAP.get(perf_score, PERF_SCORE_MAP[0])
    ws_feature["D4"] = perf_score
    # 可靠性
    ws_feature["C5"] = RELI_SCORE_MAP.get(reli_score, RELI_SCORE_MAP[0])
    ws_feature["D5"] = reli_score
    # 多重站点
    ws_feature["C6"] = MULTI_SCORE_MAP.get(multi_score, MULTI_SCORE_MAP[0])
    ws_feature["D6"] = multi_score

    # 完整性级别
    integrity = params.get("integrity_level", "C/D级别或无明确完整性级别")
    ws_feature["C7"] = integrity
    ws_feature["D7"] = INTEGRITY_FACTORS.get(integrity, 1.0)

    # 开发语言
    dev_lang = params.get("dev_language", "JAVA、C++、C#及其他同级别语言/平台")
    ws_feature["C8"] = dev_lang
    ws_feature["D8"] = LANGUAGE_FACTORS.get(dev_lang, 1.0)

    # 开发团队背景
    ws_feature["C9"] = team
    ws_feature["D9"] = TEAM_FACTORS.get(team, 1.0)

    # ── 2. 填写规模估算 ──────────────────────────
    ws_scale = wb["规模估算"]

    # C1: 规模估算方法
    method_label = "预估功能点" if is_estimated else "估算功能点"
    ws_scale["C1"] = method_label

    # 清除旧数据（第6行开始）
    for row in range(6, ws_scale.max_row + 1):
        for col in range(1, 15):
            try:
                ws_scale.cell(row=row, column=col).value = None
            except:
                pass

    # 计算功能点
    total_ufp = 0
    total_us = 0
    category_counts = {"ILF": 0, "EIF": 0, "EI": 0, "EO": 0, "EQ": 0}

    for idx, fp in enumerate(fp_items):
        row_num = 6 + idx
        cat = fp.get("category", "").upper()
        complexity = fp.get("complexity", "medium")
        reuse_level = fp.get("reuse_level", "low")
        modify_type = fp.get("modify_type", "new")

        # 获取 UFP
        if fp.get("ufp") and float(fp.get("ufp", 0)) > 0:
            ufp = float(fp["ufp"])
        else:
            ufp = _get_ufp(cat, fp_method) if cat else 0

        # 只统计有效类别
        if cat in category_counts:
            category_counts[cat] += 1

        # 计算 US (UFP × 复用率 × 修改类型)
        us = _calc_us(ufp, complexity, reuse_level, modify_type)

        total_ufp += ufp
        total_us += us

        # 写入行
        ws_scale.cell(row=row_num, column=1, value=fp.get("subsystem", ""))
        ws_scale.cell(row=row_num, column=2, value=fp.get("module_l1", ""))
        ws_scale.cell(row=row_num, column=3, value="")
        ws_scale.cell(row=row_num, column=4, value="")  # 三级模块
        ws_scale.cell(row=row_num, column=5, value="")  # 四级模块
        ws_scale.cell(row=row_num, column=6, value=fp.get("description", "")[:80])
        ws_scale.cell(row=row_num, column=7, value=fp.get("fp_name", "")[:40])
        ws_scale.cell(row=row_num, column=8, value=cat)
        ws_scale.cell(row=row_num, column=9, value=ufp)
        ws_scale.cell(row=row_num, column=10, value=REUSE_LABELS.get(reuse_level, "低"))
        ws_scale.cell(row=row_num, column=11, value=MODIFY_LABELS.get(modify_type, "新增"))
        ws_scale.cell(row=row_num, column=12, value=us)
        ws_scale.cell(row=row_num, column=13, value="")

    # 写入合计行到 M10006 (兼容公式引用)
    ws_scale["M10006"] = total_us
    ws_scale["C2"] = total_ufp
    ws_scale["C3"] = total_us * scale_factor  # 调整后功能点

    # ── 3. 填写开发费用估算（直接计算写入值） ──────
    ws_cost = wb["开发费用估算"]

    # 非功能特征因子：1 + 0.025 × Σ(各特征评分)
    nf_sum = dist_score + perf_score + reli_score + multi_score
    nf_factor = round(1 + 0.025 * nf_sum, 4)

    # 调整因子乘积
    adj_factor_product = round(app_factor * nf_factor *
                               INTEGRITY_FACTORS.get(integrity, 1.0) *
                               LANGUAGE_FACTORS.get(dev_lang, 1.0) *
                               TEAM_FACTORS.get(team, 1.0), 4)

    # 工作量计算
    pdr_lower = round(pdr * 0.8, 4)
    pdr_upper = round(pdr * 1.2, 4)
    adjusted_fp = round(total_us * scale_factor, 2)

    workload_lower = round(adjusted_fp * pdr_lower / 8, 2)
    workload_median = round(adjusted_fp * pdr / 8, 2)
    workload_upper = round(adjusted_fp * pdr_upper / 8, 2)

    # 调整后工作量
    adj_workload_lower = round(workload_lower * adj_factor_product, 2)
    adj_workload_median = round(workload_median * adj_factor_product, 2)
    adj_workload_upper = round(workload_upper * adj_factor_product, 2)

    # 人月数 = 人天 / 21.75
    pm_lower = round(adj_workload_lower / 21.75, 2)
    pm_median = round(adj_workload_median / 21.75, 2)
    pm_upper = round(adj_workload_upper / 21.75, 2)

    # 人工费（万元） = 人月 × 人月单价
    labor_lower = round(pm_lower * unit_price, 4)
    labor_median = round(pm_median * unit_price, 4)
    labor_upper = round(pm_upper * unit_price, 4)

    # 写到开发费用估算 Sheet
    ws_cost["C1"] = total_us          # 未调整功能点
    ws_cost["C2"] = scale_factor      # 规模变更调整因子
    ws_cost["C3"] = adjusted_fp       # 调整后规模
    ws_cost["C4"] = pdr_lower         # PDR下限
    ws_cost["C5"] = pdr               # PDR中值
    ws_cost["C6"] = pdr_upper         # PDR上限
    ws_cost["C7"] = workload_lower    # 未调整工作量下限(人天)
    ws_cost["C8"] = workload_median   # 未调整工作量中值(人天)
    ws_cost["C9"] = workload_upper    # 未调整工作量上限(人天)
    ws_cost["C10"] = app_factor       # 应用类型因子
    ws_cost["C11"] = nf_factor        # 非功能特征因子
    ws_cost["C12"] = INTEGRITY_FACTORS.get(integrity, 1.0)  # 完整性级别因子
    ws_cost["C13"] = LANGUAGE_FACTORS.get(dev_lang, 1.0)   # 开发语言因子
    ws_cost["C14"] = TEAM_FACTORS.get(team, 1.0)           # 团队背景因子
    ws_cost["C15"] = adj_workload_lower   # 调整后工作量下限(人天)
    ws_cost["C16"] = adj_workload_median  # 调整后工作量中值(人天)
    ws_cost["C17"] = adj_workload_upper   # 调整后工作量上限(人天)
    ws_cost["C18"] = unit_price           # 人月基准单价
    ws_cost["C19"] = labor_lower          # 基准报价下限(万元)
    ws_cost["C20"] = labor_median         # 基准报价中值(万元)
    ws_cost["C21"] = labor_upper          # 基准报价上限(万元)

    # ── 4. 保存 ──────────────────────────────────
    os.makedirs(EXPORT_DIR, exist_ok=True)
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(EXPORT_DIR, f"造价测算_{ts}.xlsx")

    wb.save(output_path)

    # ── 5. 读取已保存的值 ─────────────────────────
    wb2 = openpyxl.load_workbook(output_path, data_only=True)
    ws_cost2 = wb2["开发费用估算"]

    result = {
        "unadjusted_fp": total_us,
        "scale_factor": scale_factor,
        "adjusted_fp": adjusted_fp,
        "pdr_lower": pdr_lower,
        "pdr_median": pdr,
        "pdr_upper": pdr_upper,
        "workload_lower": workload_lower,
        "workload_median": workload_median,
        "workload_upper": workload_upper,
        "nf_factor": nf_factor,
        "app_factor": app_factor,
        "integrity_factor": INTEGRITY_FACTORS.get(integrity, 1.0),
        "language_factor": LANGUAGE_FACTORS.get(dev_lang, 1.0),
        "team_factor": TEAM_FACTORS.get(team, 1.0),
        "adj_factor_product": adj_factor_product,
        "adj_workload_lower": adj_workload_lower,
        "adj_workload_median": adj_workload_median,
        "adj_workload_upper": adj_workload_upper,
        "person_months_lower": pm_lower,
        "person_months_median": pm_median,
        "person_months_upper": pm_upper,
        "labor_cost_lower": labor_lower,
        "labor_cost_median": labor_median,
        "labor_cost_upper": labor_upper,
        "unit_price": unit_price,
        "category_counts": category_counts,
        "total_ufp": total_ufp,
        "total_us": total_us,
    }

    wb2.close()
    wb.close()

    return {"result": result, "filepath": output_path}


def read_excel_results(filepath: str) -> dict:
    """读取已生成的Excel计算结果"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["开发费用估算"]
    r = {}
    r["unadjusted_fp"] = ws["C1"].value or 0
    r["scale_factor"] = ws["C2"].value or 1
    r["adjusted_fp"] = ws["C3"].value or 0
    r["pdr_lower"] = ws["C4"].value or 0
    r["pdr_median"] = ws["C5"].value or 6.72
    r["pdr_upper"] = ws["C6"].value or 0
    r["workload_lower"] = ws["C7"].value or 0
    r["workload_median"] = ws["C8"].value or 0
    r["workload_upper"] = ws["C9"].value or 0
    r["labor_cost_median"] = ws["C20"].value or 0
    r["adjusted_fp_value"] = ws["C3"].value or 0
    wb.close()
    return r
